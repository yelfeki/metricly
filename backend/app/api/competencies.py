"""Competency Framework Database API — browsing, straw-man blueprint generation, and Excel export."""

from __future__ import annotations

import io
import json
from datetime import date, datetime, timezone
from typing import Any

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from ..core.auth import AuthUser, require_user
from ..core.database import get_db
from ..models.competency import (
    CompetencyDefinition,
    CompetencyFramework,
    CompetencyInstrumentMapping,
    CompetencyProficiencyLevel,
)
from ..models.library import Instrument, InstrumentSubscale

competencies_router = APIRouter(prefix="/competencies", tags=["competencies"])


# ---------------------------------------------------------------------------
# Pydantic response schemas
# ---------------------------------------------------------------------------


class ProficiencyLevelOut(BaseModel):
    level: int
    label: str
    descriptor: str | None = None  # free-text per-level descriptor (custom competencies)
    behavioral_indicators: list[str]
    example_behaviors: list[str]


class InstrumentMappingOut(BaseModel):
    instrument_id: str
    instrument_name: str
    instrument_short_name: str
    mapping_strength: str  # primary | supporting
    rationale: str | None
    subscale_focus: str | None
    total_items: int
    reliability_alpha: float | None
    response_format: str


class CompetencyListItem(BaseModel):
    id: str
    name: str
    definition: str | None
    framework_id: str
    framework_name: str
    factor: str | None
    cluster: str | None
    category: str | None
    role_family: str | None
    framework_source: str | None
    is_leadership: bool
    is_technical: bool
    # Custom-competency fields (see CompetencyDefinition docstring)
    is_custom: bool = False
    status: str = "active"
    organization_id: str | None = None
    created_by_user_id: str | None = None
    primary_instrument: str | None  # short name of primary instrument if any


class CompetencyDetail(BaseModel):
    id: str
    name: str
    definition: str | None
    framework_id: str
    framework_name: str
    factor: str | None
    cluster: str | None
    category: str | None
    role_family: str | None
    framework_source: str | None
    is_leadership: bool
    is_technical: bool
    # Custom-competency fields
    is_custom: bool = False
    status: str = "active"
    organization_id: str | None = None
    created_by_user_id: str | None = None
    proficiency_levels: list[ProficiencyLevelOut]
    instrument_mappings: list[InstrumentMappingOut]


# ---------------------------------------------------------------------------
# Custom competency schemas
# ---------------------------------------------------------------------------


class CustomCompetencyLevelInput(BaseModel):
    level: int  # 1..5
    label: str | None = None  # defaults to LEVEL_LABELS[level]
    descriptor: str | None = None
    behavioral_indicators: list[str] = []
    example_behaviors: list[str] = []


class CustomCompetencyCreate(BaseModel):
    name: str
    definition: str  # description
    role_family: str
    cluster: str
    framework_source: str | None = None
    levels: list[CustomCompetencyLevelInput] = []


class CustomCompetencyUpdate(BaseModel):
    name: str | None = None
    definition: str | None = None
    role_family: str | None = None
    cluster: str | None = None
    framework_source: str | None = None
    # When provided, ALL existing levels are deleted and replaced with these.
    # When None, levels are left untouched.
    levels: list[CustomCompetencyLevelInput] | None = None


class FrameworkUsageOut(BaseModel):
    competency_id: str
    framework_count: int
    framework_titles: list[str]  # capped at 5


class ClusterOptionsResponse(BaseModel):
    # role_family -> sorted unique cluster names
    options: dict[str, list[str]]


class FrameworkListItem(BaseModel):
    id: str
    name: str
    source: str | None
    description: str | None
    version: str | None
    competency_count: int


# ---------------------------------------------------------------------------
# Guided-flow ranker schemas
# ---------------------------------------------------------------------------


class RankRequest(BaseModel):
    role: str
    level: str  # "IC" | "Team Lead" | "Manager" | "Director+"
    outcome: str  # captured but not yet scored (see competency_ranker.py)
    gaps: list[str] = []  # gap-concern labels from GAP_KEYWORDS in the ranker
    size: str  # "lean" | "standard" | "comprehensive"
    required_ids: list[str] | None = None


class RankedItem(BaseModel):
    competency_id: str
    name: str
    definition: str | None
    cluster: str | None
    role_family: str | None
    framework_id: str
    framework_name: str
    score: int
    rationale: str
    suggested_proficiency_level: int


class RankResponse(BaseModel):
    role_family_inferred: str | None
    ranked: list[RankedItem]


# ---------------------------------------------------------------------------
# Straw-man schemas
# ---------------------------------------------------------------------------


class StrawManRequest(BaseModel):
    role_title: str | None = None
    initiative: str | None = None
    competency_ids: list[str] | None = None
    seniority_level: str = "mid"  # junior | mid | senior | executive
    purpose: str = "development"  # selection | development | 360


class StrawManInstrumentRef(BaseModel):
    name: str
    short_name: str
    subscale: str | None
    items: int
    alpha: float | None
    response_format: str
    rationale: str | None


class StrawManRow(BaseModel):
    competency: str
    competency_id: str
    framework: str
    factor: str | None
    cluster: str | None
    required_proficiency_level: int
    proficiency_label: str
    behavioral_indicators: list[str]
    primary_instrument: StrawManInstrumentRef | None
    supporting_instruments: list[StrawManInstrumentRef]
    assessment_method: str
    rationale: str


class StrawManResponse(BaseModel):
    title: str
    generated_at: str
    seniority_level: str
    purpose: str
    rows: list[StrawManRow]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _parse_json_field(value: str | None) -> list[str]:
    if not value:
        return []
    try:
        parsed = json.loads(value)
        return parsed if isinstance(parsed, list) else []
    except (json.JSONDecodeError, TypeError):
        return []


def _proficiency_out(pl: CompetencyProficiencyLevel) -> ProficiencyLevelOut:
    return ProficiencyLevelOut(
        level=pl.level,
        label=pl.label,
        descriptor=pl.descriptor,
        behavioral_indicators=_parse_json_field(pl.behavioral_indicators),
        example_behaviors=_parse_json_field(pl.example_behaviors),
    )


def _mapping_out(m: CompetencyInstrumentMapping, instrument: Instrument) -> InstrumentMappingOut:
    return InstrumentMappingOut(
        instrument_id=instrument.id,
        instrument_name=instrument.name,
        instrument_short_name=instrument.short_name,
        mapping_strength=m.mapping_strength,
        rationale=m.rationale,
        subscale_focus=m.subscale_focus,
        total_items=instrument.total_items,
        reliability_alpha=instrument.reliability_alpha,
        response_format=instrument.response_format,
    )


def _required_level(seniority: str, purpose: str) -> int:
    """Determine minimum required proficiency level for the blueprint."""
    base = {"junior": 1, "mid": 2, "senior": 3, "executive": 4}.get(seniority, 2)
    if purpose == "selection":
        base = min(base + 1, 5)
    return base


def _assessment_method(instrument: Instrument | None) -> str:
    if instrument is None:
        return "Structured interview / behavioural observation"
    fmt = instrument.response_format
    if "likert" in fmt:
        return "Self-report psychometric survey"
    if fmt == "forced_choice":
        return "Forced-choice psychometric survey"
    return "Multi-rater survey"


def _select_competency_ids_for_role(
    all_comps: list[CompetencyDefinition],
    seniority: str,
    purpose: str,
    role_title: str | None,
    initiative: str | None,
) -> list[str]:
    """
    Heuristic selection of competencies for a straw man when no explicit IDs are given.
    Returns up to 12 competency IDs based on role/seniority/purpose signals.
    """
    role_lower = (role_title or "").lower()
    init_lower = (initiative or "").lower()

    # ── Always-include criteria ─────────────────────────────────────────────
    always_names = {
        "Communication",
        "Teamwork and Collaboration",
        "Results Orientation",
        "Ethics and Integrity",
    }

    # ── Seniority signals ───────────────────────────────────────────────────
    if seniority in ("senior", "executive"):
        always_names.update({
            "Strategic Mindset",
            "Drives Results",
            "Develops Talent",
            "Instills Trust",
            "Leadership and Influence",
        })
    if seniority == "executive":
        always_names.update({
            "Business Insight",
            "Cultivates Innovation",
            "Builds Effective Teams",
        })

    # ── Purpose signals ─────────────────────────────────────────────────────
    if purpose == "selection":
        always_names.update({"Decision Quality", "Action Oriented", "Being Resilient"})
    if purpose == "360":
        always_names.update({
            "Communicates Effectively",
            "Emotional Intelligence",
            "Demonstrates Self-Awareness",
        })
    if purpose == "development":
        always_names.update({
            "Self-Development and Learning Agility",
            "Nimble Learning",
            "Adaptability and Flexibility",
        })

    # ── Role-keyword signals ────────────────────────────────────────────────
    sales_kw = {"sales", "account", "revenue", "commercial", "business development"}
    tech_kw = {"engineer", "developer", "software", "data", "it ", "technical", "architect"}
    people_kw = {"manager", "director", "vp", "head", "lead", "hr", "people", "talent"}
    customer_kw = {"customer", "service", "support", "success", "cx"}
    ops_kw = {"operations", "supply", "logistics", "manufacturing", "process"}

    combined_text = role_lower + " " + init_lower
    if any(k in combined_text for k in sales_kw):
        always_names.update({"Persuades", "Customer Focus", "Drives Results"})
    if any(k in combined_text for k in tech_kw):
        always_names.update({"Complex Problem Solving", "Tech Savvy", "Manages Complexity"})
    if any(k in combined_text for k in people_kw):
        always_names.update({"Drives Engagement", "Develops Talent", "Manages Conflict"})
    if any(k in combined_text for k in customer_kw):
        always_names.update({"Customer Focus", "Communication", "Service Orientation"})
    if any(k in combined_text for k in ops_kw):
        always_names.update({"Plans and Aligns", "Ensures Accountability", "Manages Complexity"})

    # Build final ID list — prefer KF + Core Behavioral, max 12
    selected: list[str] = []
    for comp in all_comps:
        if comp.name in always_names:
            selected.append(comp.id)
        if len(selected) >= 12:
            break

    # If we're short, add more from Core Behavioral
    core_beh_names = {
        "Problem Solving and Critical Thinking",
        "Adaptability and Flexibility",
        "Innovation and Creativity",
        "Diversity, Equity and Inclusion",
    }
    for comp in all_comps:
        if len(selected) >= 12:
            break
        if comp.name in core_beh_names and comp.id not in selected:
            selected.append(comp.id)

    return selected[:12]


def _build_straw_man_row(
    comp: CompetencyDefinition,
    framework_name: str,
    seniority: str,
    purpose: str,
    instrument_map: dict[str, Instrument],
) -> StrawManRow:
    req_level = _required_level(seniority, purpose)

    # Get proficiency level data
    prof_levels = sorted(comp.proficiency_levels, key=lambda x: x.level)
    # Use requested level, clamped to available
    target_pl = next(
        (pl for pl in prof_levels if pl.level == req_level),
        prof_levels[req_level - 1] if len(prof_levels) >= req_level else (prof_levels[-1] if prof_levels else None),
    )

    behavioral_indicators: list[str] = []
    if target_pl:
        behavioral_indicators = _parse_json_field(target_pl.behavioral_indicators)

    # Sort mappings: primary first
    sorted_mappings = sorted(
        comp.instrument_mappings,
        key=lambda m: (0 if m.mapping_strength == "primary" else 1),
    )

    primary_ref: StrawManInstrumentRef | None = None
    supporting_refs: list[StrawManInstrumentRef] = []

    for mapping in sorted_mappings:
        inst = instrument_map.get(mapping.instrument_id)
        if inst is None:
            continue
        ref = StrawManInstrumentRef(
            name=inst.name,
            short_name=inst.short_name,
            subscale=mapping.subscale_focus,
            items=inst.total_items,
            alpha=inst.reliability_alpha,
            response_format=inst.response_format,
            rationale=mapping.rationale,
        )
        if mapping.mapping_strength == "primary" and primary_ref is None:
            primary_ref = ref
        else:
            supporting_refs.append(ref)

    return StrawManRow(
        competency=comp.name,
        competency_id=comp.id,
        framework=framework_name,
        factor=comp.factor,
        cluster=comp.cluster,
        required_proficiency_level=req_level,
        proficiency_label=LEVEL_LABELS.get(req_level, "Proficient"),
        behavioral_indicators=behavioral_indicators[:3],  # top 3 for conciseness
        primary_instrument=primary_ref,
        supporting_instruments=supporting_refs[:2],  # top 2 supporting
        assessment_method=_assessment_method(
            instrument_map.get(sorted_mappings[0].instrument_id) if sorted_mappings else None
        ),
        rationale=(
            sorted_mappings[0].rationale if sorted_mappings else
            f"{comp.name} is a core competency relevant to this role and assessment purpose."
        ),
    )


LEVEL_LABELS = {1: "Novice", 2: "Developing", 3: "Proficient", 4: "Advanced", 5: "Expert"}


# ---------------------------------------------------------------------------
# Endpoints
# ---------------------------------------------------------------------------


@competencies_router.post("/rank-for-framework", response_model=RankResponse)
async def rank_for_framework(
    body: RankRequest,
    _: AuthUser = Depends(require_user),
    db: AsyncSession = Depends(get_db),
) -> RankResponse:
    """Read-only ranker — produces a proposal for the guided-flow screen.

    Scoring (in services/competency_ranker.py): +3 role_family match,
    +2 gap concern match, +1 cluster diversity bonus applied during greedy
    selection. Required IDs are guaranteed-included. Size cap defined by
    SIZE_CAPS (lean=6, standard=10, comprehensive=15).
    """
    from ..services.competency_ranker import infer_role_family, rank_competencies

    ranked = await rank_competencies(
        db,
        role=body.role,
        level=body.level,
        outcome=body.outcome,
        gaps=body.gaps,
        size=body.size,
        required_ids=body.required_ids,
    )
    return RankResponse(
        role_family_inferred=infer_role_family(body.role),
        ranked=[
            RankedItem(
                competency_id=r.competency_id,
                name=r.name,
                definition=r.definition,
                cluster=r.cluster,
                role_family=r.role_family,
                framework_id=r.framework_id,
                framework_name=r.framework_name,
                score=r.score,
                rationale=r.rationale,
                suggested_proficiency_level=r.suggested_proficiency_level,
            )
            for r in ranked
        ],
    )


@competencies_router.get("/frameworks", response_model=list[FrameworkListItem])
async def list_frameworks(
    _: AuthUser = Depends(require_user),
    db: AsyncSession = Depends(get_db),
) -> list[FrameworkListItem]:
    stmt = select(CompetencyFramework).options(selectinload(CompetencyFramework.competencies))
    rows = (await db.execute(stmt)).scalars().all()
    return [
        FrameworkListItem(
            id=f.id,
            name=f.name,
            source=f.source,
            description=f.description,
            version=f.version,
            competency_count=len(f.competencies),
        )
        for f in rows
    ]


@competencies_router.get("", response_model=list[CompetencyListItem])
async def list_competencies(
    framework_id: str | None = Query(None),
    factor: str | None = Query(None),
    cluster: str | None = Query(None),
    category: str | None = Query(None),
    role_family: str | None = Query(None),
    is_leadership: bool | None = Query(None),
    current_user: AuthUser = Depends(require_user),
    db: AsyncSession = Depends(get_db),
) -> list[CompetencyListItem]:
    from ..services.competency_visibility import visible_competencies_stmt

    stmt = visible_competencies_stmt(current_user.user_id).options(
        selectinload(CompetencyDefinition.framework),
        selectinload(CompetencyDefinition.instrument_mappings),
    )
    if framework_id:
        stmt = stmt.where(CompetencyDefinition.framework_id == framework_id)
    if factor:
        stmt = stmt.where(CompetencyDefinition.factor == factor)
    if cluster:
        stmt = stmt.where(CompetencyDefinition.cluster == cluster)
    if category:
        stmt = stmt.where(CompetencyDefinition.category == category)
    if role_family:
        stmt = stmt.where(CompetencyDefinition.role_family == role_family)
    if is_leadership is not None:
        stmt = stmt.where(CompetencyDefinition.is_leadership == is_leadership)

    rows = (await db.execute(stmt)).scalars().all()

    # Find primary instrument short names
    inst_ids = {
        m.instrument_id
        for c in rows
        for m in c.instrument_mappings
        if m.mapping_strength == "primary"
    }
    insts: dict[str, str] = {}
    if inst_ids:
        inst_rows = (
            await db.execute(select(Instrument).where(Instrument.id.in_(inst_ids)))
        ).scalars().all()
        insts = {i.id: i.short_name for i in inst_rows}

    return [
        CompetencyListItem(
            id=c.id,
            name=c.name,
            definition=c.definition,
            framework_id=c.framework_id,
            framework_name=c.framework.name if c.framework else "",
            factor=c.factor,
            cluster=c.cluster,
            category=c.category,
            role_family=c.role_family,
            framework_source=c.framework_source,
            is_leadership=c.is_leadership,
            is_technical=c.is_technical,
            is_custom=c.is_custom,
            status=c.status,
            organization_id=c.organization_id,
            created_by_user_id=c.created_by_user_id,
            primary_instrument=next(
                (insts[m.instrument_id] for m in c.instrument_mappings
                 if m.mapping_strength == "primary" and m.instrument_id in insts),
                None,
            ),
        )
        for c in rows
    ]


async def _build_competency_detail(
    db: AsyncSession, comp: CompetencyDefinition
) -> CompetencyDetail:
    """Shape a CompetencyDefinition into the API response (loads instruments)."""
    inst_ids = [m.instrument_id for m in comp.instrument_mappings]
    insts: dict[str, Instrument] = {}
    if inst_ids:
        inst_rows = (
            await db.execute(select(Instrument).where(Instrument.id.in_(inst_ids)))
        ).scalars().all()
        insts = {i.id: i for i in inst_rows}

    return CompetencyDetail(
        id=comp.id,
        name=comp.name,
        definition=comp.definition,
        framework_id=comp.framework_id,
        framework_name=comp.framework.name if comp.framework else "",
        factor=comp.factor,
        cluster=comp.cluster,
        category=comp.category,
        role_family=comp.role_family,
        framework_source=comp.framework_source,
        is_leadership=comp.is_leadership,
        is_technical=comp.is_technical,
        is_custom=comp.is_custom,
        status=comp.status,
        organization_id=comp.organization_id,
        created_by_user_id=comp.created_by_user_id,
        proficiency_levels=[
            _proficiency_out(pl) for pl in sorted(comp.proficiency_levels, key=lambda p: p.level)
        ],
        instrument_mappings=[
            _mapping_out(m, insts[m.instrument_id])
            for m in comp.instrument_mappings
            if m.instrument_id in insts
        ],
    )


# NOTE: keep `/cluster-options` BEFORE `/{competency_id}` — FastAPI matches
# routes in declaration order, and "cluster-options" would otherwise be
# captured as a competency_id path parameter and return 404.
@competencies_router.get("/cluster-options", response_model=ClusterOptionsResponse)
async def get_cluster_options(
    current_user: AuthUser = Depends(require_user),
    db: AsyncSession = Depends(get_db),
) -> ClusterOptionsResponse:
    """Return role_family → sorted unique cluster names visible to this user.

    Visibility helper applied so a user sees seeded clusters plus any their
    org has introduced via custom competencies.
    """
    from ..services.competency_visibility import visible_competencies_stmt

    rows = (
        await db.execute(visible_competencies_stmt(current_user.user_id))
    ).scalars().all()

    grouped: dict[str, set[str]] = {}
    for c in rows:
        if not c.role_family or not c.cluster:
            continue
        grouped.setdefault(c.role_family, set()).add(c.cluster)

    return ClusterOptionsResponse(
        options={rf: sorted(clusters) for rf, clusters in sorted(grouped.items())}
    )


@competencies_router.get("/{competency_id}", response_model=CompetencyDetail)
async def get_competency(
    competency_id: str,
    current_user: AuthUser = Depends(require_user),
    db: AsyncSession = Depends(get_db),
) -> CompetencyDetail:
    from ..services.competency_visibility import visible_competencies_stmt

    stmt = (
        visible_competencies_stmt(current_user.user_id)
        .where(CompetencyDefinition.id == competency_id)
        .options(
            selectinload(CompetencyDefinition.framework),
            selectinload(CompetencyDefinition.proficiency_levels),
            selectinload(CompetencyDefinition.instrument_mappings),
        )
    )
    comp = (await db.execute(stmt)).scalar_one_or_none()
    if comp is None:
        # Either doesn't exist or isn't visible to this user — same response.
        raise HTTPException(status_code=404, detail="Competency not found.")
    return await _build_competency_detail(db, comp)


# ---------------------------------------------------------------------------
# Custom-competency endpoints
# ---------------------------------------------------------------------------

_CUSTOM_FRAMEWORK_NAME = "Custom Competencies"


async def _get_or_create_custom_framework(db: AsyncSession) -> CompetencyFramework:
    """Lazy-init the single global parent framework for all custom competencies.

    Custom rows still need a non-null framework_id (the existing schema), but
    the org boundary lives on the row, not on the parent. One row in
    competency_frameworks named "Custom Competencies" serves as the parent for
    every org's custom rows.
    """
    fw = (
        await db.execute(
            select(CompetencyFramework).where(
                CompetencyFramework.name == _CUSTOM_FRAMEWORK_NAME
            )
        )
    ).scalar_one_or_none()
    if fw is not None:
        return fw
    fw = CompetencyFramework(
        name=_CUSTOM_FRAMEWORK_NAME,
        source="Custom (org-created)",
        description=(
            "Parent framework for organisation-created competencies. Org scope is "
            "enforced on each CompetencyDefinition.organization_id; this framework "
            "is shared across all orgs as a logical container."
        ),
        version="1.0",
    )
    db.add(fw)
    await db.flush()
    return fw


def _level_has_indicators(indicators_json: str | None) -> bool:
    parsed = _parse_json_field(indicators_json)
    return any(isinstance(s, str) and s.strip() for s in parsed)


async def _compute_status_from_db(db: AsyncSession, competency_id: str) -> str:
    """Re-derive status from the persisted level rows. Call after writes."""
    from ..services.competency_visibility import derive_status

    levels = (
        await db.execute(
            select(CompetencyProficiencyLevel).where(
                CompetencyProficiencyLevel.competency_id == competency_id
            )
        )
    ).scalars().all()
    level_count = len(levels)
    with_inds = sum(1 for lv in levels if _level_has_indicators(lv.behavioral_indicators))
    return derive_status(level_count=level_count, levels_with_indicators=with_inds)


async def _write_levels(
    db: AsyncSession,
    competency_id: str,
    inputs: list[CustomCompetencyLevelInput],
) -> None:
    """Insert proficiency level rows. Caller must delete prior rows first if updating."""
    for lvl in inputs:
        db.add(
            CompetencyProficiencyLevel(
                competency_id=competency_id,
                level=lvl.level,
                label=lvl.label or LEVEL_LABELS.get(lvl.level, f"Level {lvl.level}"),
                descriptor=lvl.descriptor,
                behavioral_indicators=json.dumps(lvl.behavioral_indicators),
                example_behaviors=json.dumps(lvl.example_behaviors),
            )
        )


@competencies_router.post("/custom", response_model=CompetencyDetail, status_code=201)
async def create_custom_competency(
    body: CustomCompetencyCreate,
    current_user: AuthUser = Depends(require_user),
    db: AsyncSession = Depends(get_db),
) -> CompetencyDetail:
    """Create an organisation-scoped custom competency.

    Sets organization_id = created_by_user_id = current_user.user_id (until the
    Organization model ships, see services.competency_visibility). Computes
    status from the supplied levels via derive_status — clients cannot lie
    about completeness.
    """
    fw = await _get_or_create_custom_framework(db)
    comp = CompetencyDefinition(
        framework_id=fw.id,
        name=body.name,
        definition=body.definition,
        role_family=body.role_family,
        cluster=body.cluster,
        framework_source=body.framework_source,
        organization_id=current_user.user_id,
        created_by_user_id=current_user.user_id,
        is_custom=True,
        status="draft",  # provisional — recomputed below
    )
    db.add(comp)
    await db.flush()  # get comp.id

    if body.levels:
        await _write_levels(db, comp.id, body.levels)
        await db.flush()

    comp.status = await _compute_status_from_db(db, comp.id)
    await db.commit()

    # Reload with relationships for response
    refreshed = (
        await db.execute(
            select(CompetencyDefinition)
            .where(CompetencyDefinition.id == comp.id)
            .options(
                selectinload(CompetencyDefinition.framework),
                selectinload(CompetencyDefinition.proficiency_levels),
                selectinload(CompetencyDefinition.instrument_mappings),
            )
        )
    ).scalar_one()
    return await _build_competency_detail(db, refreshed)


@competencies_router.patch("/custom/{competency_id}", response_model=CompetencyDetail)
async def update_custom_competency(
    competency_id: str,
    body: CustomCompetencyUpdate,
    current_user: AuthUser = Depends(require_user),
    db: AsyncSession = Depends(get_db),
) -> CompetencyDetail:
    """Update a custom competency. 403 unless is_editable_by(comp, user)."""
    from ..services.competency_visibility import is_editable_by

    comp = (
        await db.execute(
            select(CompetencyDefinition)
            .where(CompetencyDefinition.id == competency_id)
            .options(
                selectinload(CompetencyDefinition.proficiency_levels),
            )
        )
    ).scalar_one_or_none()
    if comp is None:
        raise HTTPException(status_code=404, detail="Competency not found.")
    if not is_editable_by(comp, current_user.user_id):
        raise HTTPException(
            status_code=403,
            detail="Only custom competencies owned by your organisation can be edited.",
        )

    if body.name is not None:
        comp.name = body.name
    if body.definition is not None:
        comp.definition = body.definition
    if body.role_family is not None:
        comp.role_family = body.role_family
    if body.cluster is not None:
        comp.cluster = body.cluster
    if body.framework_source is not None:
        comp.framework_source = body.framework_source

    if body.levels is not None:
        # Replace all existing levels with the new set (atomic semantic)
        for existing in list(comp.proficiency_levels):
            await db.delete(existing)
        await db.flush()
        await _write_levels(db, comp.id, body.levels)
        await db.flush()

    comp.status = await _compute_status_from_db(db, comp.id)
    await db.commit()

    # populate_existing=True forces the relationship cache on the identity-mapped
    # comp object to be rebuilt — otherwise selectinload skips the refresh and
    # returns whatever was loaded at the start of the request.
    refreshed = (
        await db.execute(
            select(CompetencyDefinition)
            .where(CompetencyDefinition.id == comp.id)
            .execution_options(populate_existing=True)
            .options(
                selectinload(CompetencyDefinition.framework),
                selectinload(CompetencyDefinition.proficiency_levels),
                selectinload(CompetencyDefinition.instrument_mappings),
            )
        )
    ).scalar_one()
    return await _build_competency_detail(db, refreshed)


@competencies_router.get(
    "/custom/{competency_id}/framework-usage", response_model=FrameworkUsageOut
)
async def get_framework_usage(
    competency_id: str,
    current_user: AuthUser = Depends(require_user),
    db: AsyncSession = Depends(get_db),
) -> FrameworkUsageOut:
    """How many frameworks import this competency? Drives the edit-confirmation modal."""
    from ..models.framework import Competency as UserCompetency, Framework as UserFramework
    from ..services.competency_visibility import visible_competencies_stmt

    # Visibility check — only the competency's owner (or seeded competencies) can read usage.
    visible = (
        await db.execute(
            visible_competencies_stmt(current_user.user_id).where(
                CompetencyDefinition.id == competency_id
            )
        )
    ).scalar_one_or_none()
    if visible is None:
        raise HTTPException(status_code=404, detail="Competency not found.")

    stmt = (
        select(UserFramework.title)
        .join(UserCompetency, UserCompetency.framework_id == UserFramework.id)
        .where(UserCompetency.library_competency_id == competency_id)
        .where(UserFramework.user_id == current_user.user_id)
    )
    titles = [row[0] for row in (await db.execute(stmt)).all()]
    return FrameworkUsageOut(
        competency_id=competency_id,
        framework_count=len(titles),
        framework_titles=titles[:5],
    )


@competencies_router.post("/straw-man", response_model=StrawManResponse)
async def generate_straw_man(
    body: StrawManRequest,
    _: AuthUser = Depends(require_user),
    db: AsyncSession = Depends(get_db),
) -> StrawManResponse:
    """Generate an assessment blueprint straw man."""

    # Load competencies
    if body.competency_ids:
        stmt = (
            select(CompetencyDefinition)
            .where(CompetencyDefinition.id.in_(body.competency_ids))
            .options(
                selectinload(CompetencyDefinition.framework),
                selectinload(CompetencyDefinition.proficiency_levels),
                selectinload(CompetencyDefinition.instrument_mappings),
            )
        )
        comps = (await db.execute(stmt)).scalars().all()
    else:
        # Heuristic selection
        all_comps_stmt = (
            select(CompetencyDefinition)
            .options(
                selectinload(CompetencyDefinition.framework),
                selectinload(CompetencyDefinition.proficiency_levels),
                selectinload(CompetencyDefinition.instrument_mappings),
            )
        )
        all_comps = (await db.execute(all_comps_stmt)).scalars().all()
        selected_ids = _select_competency_ids_for_role(
            all_comps, body.seniority_level, body.purpose, body.role_title, body.initiative
        )
        comps = [c for c in all_comps if c.id in set(selected_ids)]

    if not comps:
        raise HTTPException(status_code=404, detail="No competencies found for the given criteria.")

    # Load all referenced instruments
    inst_ids = {m.instrument_id for c in comps for m in c.instrument_mappings}
    insts: dict[str, Instrument] = {}
    if inst_ids:
        inst_rows = (
            await db.execute(select(Instrument).where(Instrument.id.in_(inst_ids)))
        ).scalars().all()
        insts = {i.id: i for i in inst_rows}

    # Load framework names
    fw_ids = {c.framework_id for c in comps}
    fw_rows = (
        await db.execute(select(CompetencyFramework).where(CompetencyFramework.id.in_(fw_ids)))
    ).scalars().all()
    fw_names: dict[str, str] = {f.id: f.name for f in fw_rows}

    rows = [
        _build_straw_man_row(
            comp=c,
            framework_name=fw_names.get(c.framework_id, ""),
            seniority=body.seniority_level,
            purpose=body.purpose,
            instrument_map=insts,
        )
        for c in comps
    ]

    role_part = body.role_title or body.initiative or "Assessment"
    title = f"Assessment Blueprint: {role_part}"

    return StrawManResponse(
        title=title,
        generated_at=datetime.now(timezone.utc).isoformat(),
        seniority_level=body.seniority_level,
        purpose=body.purpose,
        rows=rows,
    )


@competencies_router.post("/straw-man/export")
async def export_straw_man(
    body: StrawManRequest,
    _: AuthUser = Depends(require_user),
    db: AsyncSession = Depends(get_db),
) -> StreamingResponse:
    """Generate and return an Excel (.xlsx) assessment blueprint."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import (
            Alignment,
            Border,
            Font,
            PatternFill,
            Side,
        )
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise HTTPException(
            status_code=503,
            detail="openpyxl not installed. Run: pip install openpyxl",
        ) from exc

    # Re-use the straw-man logic
    straw_man = await generate_straw_man(body, _, db)

    wb = Workbook()
    ws = wb.active
    ws.title = "Assessment Blueprint"

    # ── Factor colour palette ────────────────────────────────────────────────
    FACTOR_COLORS: dict[str, str] = {
        "Thought": "DBEAFE",   # blue-100
        "Results": "D1FAE5",   # green-100
        "People": "EDE9FE",    # violet-100
        "Self": "FEF3C7",      # amber-100
        "Content Skills": "FEE2E2",
        "Process Skills": "FEF9C3",
        "Social Skills": "DCFCE7",
        "Complex Problem Solving": "E0F2FE",
        "Technical Skills": "F3E8FF",
        "Systems Skills": "FFE4E6",
        "Resource Management": "F0FDF4",
        "default": "F9FAFB",   # gray-50
    }

    def _factor_color(factor: str | None) -> str:
        return FACTOR_COLORS.get(factor or "default", FACTOR_COLORS["default"])

    # ── Header row ───────────────────────────────────────────────────────────
    headers = [
        "Competency", "Framework", "Factor", "Cluster",
        "Required Proficiency", "Behavioral Indicators",
        "Primary Measure", "Subscale", "Items", "α",
        "Supporting Measures", "Assessment Method", "Rationale",
    ]

    header_fill = PatternFill(start_color="1E1B4B", end_color="1E1B4B", fill_type="solid")
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    thin = Side(border_style="thin", color="D1D5DB")
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap_align = Alignment(wrap_text=True, vertical="top")
    center_align = Alignment(horizontal="center", vertical="top")

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = cell_border

    # ── Data rows ────────────────────────────────────────────────────────────
    for row_idx, row in enumerate(straw_man.rows, start=2):
        fill_color = _factor_color(row.factor)
        row_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

        supporting = "; ".join(
            f"{s.short_name}" + (f" ({s.subscale})" if s.subscale else "")
            for s in row.supporting_instruments
        )
        indicators_text = "\n".join(f"• {ind}" for ind in row.behavioral_indicators)

        values = [
            row.competency,
            row.framework,
            row.factor or "",
            row.cluster or "",
            f"{row.required_proficiency_level} — {row.proficiency_label}",
            indicators_text,
            row.primary_instrument.short_name if row.primary_instrument else "Behavioural Interview",
            row.primary_instrument.subscale if row.primary_instrument and row.primary_instrument.subscale else "",
            row.primary_instrument.items if row.primary_instrument else "",
            round(row.primary_instrument.alpha, 2) if row.primary_instrument and row.primary_instrument.alpha else "",
            supporting,
            row.assessment_method,
            row.rationale or "",
        ]

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = row_fill
            cell.border = cell_border
            cell.alignment = wrap_align if col_idx in (6, 11, 13) else (
                center_align if col_idx in (5, 9, 10) else
                Alignment(vertical="top")
            )
            cell.font = Font(name="Calibri", size=10)

    # ── Column widths ────────────────────────────────────────────────────────
    col_widths = [22, 20, 14, 18, 18, 38, 14, 16, 7, 6, 24, 22, 40]
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── Freeze top row ───────────────────────────────────────────────────────
    ws.freeze_panes = "A2"

    # ── Metadata sheet ───────────────────────────────────────────────────────
    meta_ws = wb.create_sheet("Metadata")
    meta_ws.append(["Generated by", "Metricly Assessment Platform"])
    meta_ws.append(["Generated at", straw_man.generated_at])
    meta_ws.append(["Role / Initiative", body.role_title or body.initiative or ""])
    meta_ws.append(["Seniority Level", body.seniority_level])
    meta_ws.append(["Assessment Purpose", body.purpose])
    meta_ws.append(["Total Competencies", len(straw_man.rows)])

    # ── Write to buffer ──────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    role_slug = (body.role_title or body.initiative or "Blueprint").replace(" ", "_")[:30]
    today = date.today().isoformat()
    filename = f"Metricly_Assessment_Blueprint_{role_slug}_{today}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
