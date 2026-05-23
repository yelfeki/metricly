"""
Survey import — template download + file parse.

Routes (registered with prefix /api/v1):
  GET  /surveys/import/template  →  .xlsx download
  POST /surveys/import           →  ImportResult JSON
"""
import io
from typing import Literal

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from ..core.auth import AuthUser, require_user

import_router = APIRouter(prefix="/surveys", tags=["surveys"])

# ---------------------------------------------------------------------------
# Column metadata
# ---------------------------------------------------------------------------

HEADERS = [
    "Question #",
    "Subscale / Section",
    "Question Text",
    "Question Type",
    "Option 1",
    "Option 2",
    "Option 3",
    "Option 4",
    "Option 5",
    "Option 6",
    "Reverse Scored?",
    "Required?",
    "Notes",
]

INSTRUCTIONS_ROW = [
    "Auto-filled",
    "Optional — e.g. Vigor, Dedication, Section 1",
    "Required — type your question text here",
    "Choose from dropdown",
    "For MC / Forced Choice — blank for Likert/text",
    "Option 2 (leave blank if fewer than 6)",
    "Option 3",
    "Option 4",
    "Option 5",
    "Option 6",
    "Yes / No — reverse-score this item?",
    "Yes / No — is this required? (default Yes)",
    "Internal notes — not shown to respondents",
]

QUESTION_TYPES = [
    "Likert 1-5",
    "Likert 1-7",
    "Likert 1-10",
    "Multiple Choice (Single)",
    "Multiple Choice (Multiple)",
    "Forced Choice (Pick 1 of 2)",
    "Yes / No",
    "True / False",
    "Open Text",
    "Rating Scale",
]

TYPE_MAP: dict[str, dict] = {
    "Likert 1-5":                  {"question_type": "likert", "likert_min": 1, "likert_max": 5},
    "Likert 1-7":                  {"question_type": "likert", "likert_min": 1, "likert_max": 7},
    "Likert 1-10":                 {"question_type": "likert", "likert_min": 1, "likert_max": 10},
    "Multiple Choice (Single)":    {"question_type": "single_choice"},
    "Multiple Choice (Multiple)":  {"question_type": "multiple_choice"},
    "Forced Choice (Pick 1 of 2)": {"question_type": "forced_choice"},
    "Yes / No":                    {"question_type": "yes_no"},
    "True / False":                {"question_type": "true_false"},
    "Open Text":                   {"question_type": "text"},
    "Rating Scale":                {"question_type": "rating"},
}

# ---------------------------------------------------------------------------
# Schemas
# ---------------------------------------------------------------------------


class ImportedQuestion(BaseModel):
    text: str
    question_type: Literal[
        "likert", "single_choice", "multiple_choice",
        "forced_choice", "yes_no", "true_false", "text", "rating"
    ]
    likert_min: int | None = None
    likert_max: int | None = None
    options: list[str] = []
    reverse_scored: bool = False
    required: bool = True
    subscale: str | None = None
    notes: str | None = None


class ImportResult(BaseModel):
    source: Literal["template", "unstructured"]
    questions: list[ImportedQuestion]


# ---------------------------------------------------------------------------
# GET /surveys/import/template
# ---------------------------------------------------------------------------


@import_router.get("/import/template")
async def download_import_template(user: AuthUser = Depends(require_user)):
    """Generate and return a formatted .xlsx survey import template."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = Workbook()

    # -----------------------------------------------------------------------
    # Sheet 1: Survey Questions
    # -----------------------------------------------------------------------
    ws = wb.active
    ws.title = "Survey Questions"

    header_fill = PatternFill("solid", fgColor="4C1D95")
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    instr_font = Font(name="Calibri", italic=True, color="6B7280", size=10)
    instr_fill = PatternFill("solid", fgColor="F5F3FF")
    row_fill_a = PatternFill("solid", fgColor="FAF8FF")
    row_fill_b = PatternFill("solid", fgColor="FFFFFF")
    num_font = Font(name="Calibri", color="9CA3AF", size=10)
    center = Alignment(horizontal="center", vertical="center")
    left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Row 1: bold headers
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    # Row 2: instruction row
    for col_idx, instr in enumerate(INSTRUCTIONS_ROW, start=1):
        cell = ws.cell(row=2, column=col_idx, value=instr)
        cell.fill = instr_fill
        cell.font = instr_font
        cell.alignment = left_wrap

    # Rows 3-52: 50 blank question rows, alternating tint
    for row_num in range(3, 53):
        q_num = row_num - 2
        fill = row_fill_a if q_num % 2 == 1 else row_fill_b

        a_cell = ws.cell(row=row_num, column=1, value=q_num)
        a_cell.fill = fill
        a_cell.font = num_font
        a_cell.alignment = center

        for col_idx in range(2, len(HEADERS) + 1):
            ws.cell(row=row_num, column=col_idx).fill = fill

    # Data validation — Col D: Question Type
    dv_type = DataValidation(
        type="list",
        formula1='"' + ",".join(QUESTION_TYPES) + '"',
        allow_blank=True,
        showDropDown=False,
    )
    dv_type.sqref = "D3:D52"
    ws.add_data_validation(dv_type)

    # Data validation — Col K: Reverse Scored
    dv_rev = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True, showDropDown=False)
    dv_rev.sqref = "K3:K52"
    ws.add_data_validation(dv_rev)

    # Data validation — Col L: Required
    dv_req = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True, showDropDown=False)
    dv_req.sqref = "L3:L52"
    ws.add_data_validation(dv_req)

    # Column widths (A-M)
    col_widths = [6, 22, 60, 28, 16, 16, 16, 16, 16, 16, 16, 12, 28]
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Row heights
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 36
    for r in range(3, 53):
        ws.row_dimensions[r].height = 18

    # Freeze top 2 rows
    ws.freeze_panes = "A3"

    # -----------------------------------------------------------------------
    # Sheet 2: Instructions
    # -----------------------------------------------------------------------
    wi = wb.create_sheet("Instructions")
    wi.column_dimensions["A"].width = 28
    wi.column_dimensions["B"].width = 72

    title_font = Font(name="Calibri", bold=True, size=14, color="4C1D95")
    section_fill = PatternFill("solid", fgColor="EDE9FE")
    heading_font = Font(name="Calibri", bold=True, size=10, color="1E1B4B")
    body_font = Font(name="Calibri", size=10, color="374151")
    example_font = Font(name="Calibri", size=10, color="6B7280", italic=True)
    wrap = Alignment(wrap_text=True, vertical="top")

    wi.cell(row=1, column=1, value="Metricly Survey Template — Instructions").font = title_font
    wi.merge_cells("A1:B1")
    wi.row_dimensions[1].height = 28

    content_rows = [
        ("", ""),
        ("COLUMN GUIDE", None),
        ("Question #",           "Auto-filled (1–50). Do not edit this column."),
        ("Subscale / Section",   "Optional grouping label — e.g. 'Vigor', 'Dedication', 'Section 1'. Used to organise questions into factors/subscales."),
        ("Question Text",        "REQUIRED. The full question text as it will appear to respondents."),
        ("Question Type",        "Choose from the dropdown list. See type guide below."),
        ("Option 1–6",           "For Multiple Choice and Forced Choice types only. Leave blank for Likert, Open Text, Yes/No, True/False, and Rating."),
        ("Reverse Scored?",      "Yes or No. Applies only to Likert items. Reverses the scoring direction (e.g. 5→1 for a 1–5 scale)."),
        ("Required?",            "Yes or No. Whether the respondent must answer this question. Default: Yes."),
        ("Notes",                "Internal notes — not shown to respondents."),
        ("", ""),
        ("QUESTION TYPE GUIDE", None),
        ("Likert 1-5",           "5-point agreement scale (Strongly Disagree to Strongly Agree). Leave Option columns blank."),
        ("",                     "Example: 'I feel motivated at work.' — no options needed."),
        ("Likert 1-7",           "7-point agreement scale. Leave Option columns blank."),
        ("Likert 1-10",          "10-point scale. Leave Option columns blank."),
        ("Multiple Choice (Single)", "Respondent selects exactly one option. Fill Option 1–6 with your choices."),
        ("",                     "Example: 'Which department are you in?' → Option 1: HR | Option 2: Engineering | Option 3: Sales"),
        ("Multiple Choice (Multiple)", "Respondent may select multiple options. Fill Option 1–6 with your choices."),
        ("",                     "Example: 'Select all that apply:' → Option 1: Remote | Option 2: Hybrid | Option 3: On-site"),
        ("Forced Choice (Pick 1 of 2)", "Respondent picks one of exactly 2 options. Fill Option 1 and Option 2 only."),
        ("",                     "Example: 'I prefer working...' → Option 1: Alone | Option 2: In teams"),
        ("Yes / No",             "Simple Yes/No question. Leave Option columns blank — options are auto-filled."),
        ("True / False",         "Simple True/False question. Leave Option columns blank — options are auto-filled."),
        ("Open Text",            "Free-text response box. Leave Option columns blank."),
        ("Rating Scale",         "Numeric rating (1–5). Leave Option columns blank."),
        ("", ""),
        ("TIPS", None),
        ("",                     "• Save as .xlsx (Excel format) before uploading."),
        ("",                     "• Up to 50 questions per import."),
        ("",                     "• Start filling from row 3 — rows 1 and 2 are reserved."),
        ("",                     "• Empty rows are automatically skipped."),
        ("",                     "• After import you can edit question types and options in the review step."),
    ]

    row = 2
    for label, value in content_rows:
        if value is None:
            # Section heading
            cell = wi.cell(row=row, column=1, value=label)
            cell.font = Font(name="Calibri", bold=True, size=11, color="4C1D95")
            cell.fill = section_fill
            wi.cell(row=row, column=2).fill = section_fill
            wi.merge_cells(f"A{row}:B{row}")
            wi.row_dimensions[row].height = 20
        elif label:
            a = wi.cell(row=row, column=1, value=label)
            a.font = heading_font
            b = wi.cell(row=row, column=2, value=value)
            is_example = value.startswith("Example")
            b.font = example_font if is_example else body_font
            b.alignment = wrap
        else:
            b = wi.cell(row=row, column=2, value=value)
            is_example = value.startswith("Example") or value.startswith("•")
            b.font = example_font if value.startswith("Example") else body_font
            b.alignment = wrap
        row += 1

    # Stream response
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="Metricly_Survey_Template.xlsx"'},
    )


# ---------------------------------------------------------------------------
# Template detection + parsing
# ---------------------------------------------------------------------------


def _is_template(ws) -> bool:
    """Return True if the first row matches the Metricly template header exactly."""
    expected = [h.lower() for h in HEADERS]
    actual = [
        str(ws.cell(row=1, column=c).value or "").strip().lower()
        for c in range(1, len(HEADERS) + 1)
    ]
    return actual == expected


def _parse_template(ws) -> list[ImportedQuestion]:
    """Parse questions from a Metricly-formatted worksheet (rows 3+)."""
    questions: list[ImportedQuestion] = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        # Skip fully empty rows (check cols B-M)
        if not any(row[i] for i in range(1, 13)):
            continue

        text = str(row[2] or "").strip()
        if not text:
            continue

        subscale = str(row[1] or "").strip() or None
        qtype_raw = str(row[3] or "").strip()
        options = [str(row[i] or "").strip() for i in range(4, 10)]
        options = [o for o in options if o]
        reverse = str(row[10] or "").strip().lower() == "yes"
        required = str(row[11] or "").strip().lower() != "no"
        notes = str(row[12] or "").strip() or None

        mapping = TYPE_MAP.get(qtype_raw, {"question_type": "text"})
        questions.append(ImportedQuestion(
            text=text,
            subscale=subscale,
            options=options,
            reverse_scored=reverse,
            required=required,
            notes=notes,
            **mapping,
        ))
    return questions


def _parse_unstructured(ws) -> list[ImportedQuestion]:
    """Best-effort heuristic parse of a generic (non-template) Excel worksheet."""
    questions: list[ImportedQuestion] = []
    for row in ws.iter_rows(min_row=1, values_only=True):
        text_cells = [str(c).strip() for c in row if c is not None and str(c).strip()]
        if not text_cells:
            continue
        candidate = text_cells[0]
        # Heuristic: looks like a question if it has enough text
        if len(candidate) >= 10:
            questions.append(ImportedQuestion(
                text=candidate,
                question_type="text",
            ))
    return questions


# ---------------------------------------------------------------------------
# POST /surveys/import
# ---------------------------------------------------------------------------


_ALLOWED_EXTENSIONS = (".xlsx", ".xls")
_ALLOWED_MIME = {
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.ms-excel",
    "application/octet-stream",
}


@import_router.post("/import", response_model=ImportResult)
async def import_survey_file(
    file: UploadFile = File(...),
    user: AuthUser = Depends(require_user),
):
    """
    Parse an uploaded Excel (.xlsx) file.
    - Metricly template format: detected by column headers, parsed precisely.
    - Generic Excel: best-effort heuristic extraction.
    """
    filename = file.filename or ""
    is_excel = (
        file.content_type in _ALLOWED_MIME
        or any(filename.lower().endswith(ext) for ext in _ALLOWED_EXTENSIONS)
    )
    if not is_excel:
        raise HTTPException(
            status_code=415,
            detail=(
                "Only Excel (.xlsx) files are supported. "
                "Download the Metricly template, fill it in, and upload it here."
            ),
        )

    content = await file.read()
    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(
            status_code=422,
            detail="Could not read file. Please ensure it is a valid .xlsx file.",
        )

    ws = wb.active

    if _is_template(ws):
        questions = _parse_template(ws)
        source: Literal["template", "unstructured"] = "template"
    else:
        questions = _parse_unstructured(ws)
        source = "unstructured"

    if not questions:
        raise HTTPException(
            status_code=422,
            detail=(
                "No questions found. Fill in Column C (Question Text) "
                "starting from row 3, or use the Metricly template."
            ),
        )

    return ImportResult(source=source, questions=questions)
